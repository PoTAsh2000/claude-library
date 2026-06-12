#!/usr/bin/env python3
"""
Read an Excel file and output its contents as a markdown string to stdout.
Usage: read_excel.py <file_path> [sheet_name]
"""
import os
import re
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding="utf-8")

MAX_ROWS = 500

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell as OpenpyxlMergedCell
except ImportError:
    print("Error: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def normalize_path(p):
    """Convert Git Bash POSIX path /c/Users/... to C:/Users/... for Windows Python."""
    if not p:
        return p
    m = re.match(r"^/([a-zA-Z])/(.*)", p)
    if m:
        return m.group(1).upper() + ":/" + m.group(2)
    return p


def format_value(val):
    """Format a raw cell value for a markdown table cell."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        s = str(int(val)) if val == int(val) else str(val)
    else:
        s = str(val).strip()
    # Escape order matters: backslash first, then pipe, then inline markers
    s = s.replace("\\", "\\\\")
    s = s.replace("|", "\\|")
    s = s.replace("\n", " ").replace("\r", "")
    for ch in ("*", "_", "`", "#"):
        s = s.replace(ch, "\\" + ch)
    return s


def build_merge_map(ws):
    """Return {(row, col): top_left_value} for all non-top-left cells in merged regions."""
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        tl_row, tl_col = merge_range.min_row, merge_range.min_col
        tl_value = ws.cell(tl_row, tl_col).value
        for row_cells in merge_range.rows:
            for (r, c) in row_cells:
                if (r, c) != (tl_row, tl_col):
                    merge_map[(r, c)] = tl_value
    return merge_map


def get_cell_value(cell, merge_map):
    if isinstance(cell, OpenpyxlMergedCell):
        return merge_map.get((cell.row, cell.column))
    return cell.value


def sheet_to_markdown(ws):
    merge_map = build_merge_map(ws)
    rows = list(ws.iter_rows())

    # Skip leading empty rows
    start = next(
        (i for i, row in enumerate(rows) if any(get_cell_value(c, merge_map) is not None for c in row)),
        None,
    )
    if start is None:
        return "*Sheet is empty.*"

    rows = rows[start:]
    headers = [format_value(get_cell_value(c, merge_map)) or f"Column {i + 1}" for i, c in enumerate(rows[0])]

    while headers and not headers[-1].strip():
        headers.pop()

    if not headers:
        return "*Sheet is empty.*"

    ncols = len(headers)
    data_rows = rows[1:]
    truncated = len(data_rows) > MAX_ROWS
    if truncated:
        data_rows = data_rows[:MAX_ROWS]

    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * ncols) + " |",
    ]

    for row in data_rows:
        cells = [format_value(get_cell_value(c, merge_map)) for c in row[:ncols]]
        while len(cells) < ncols:
            cells.append("")
        lines.append("| " + " | ".join(cells) + " |")

    result = "\n".join(lines)
    if truncated:
        result += f"\n\n> **Note:** Output truncated to {MAX_ROWS} rows. The sheet contains more data."
    return result


def read_xlsx(file_path, target_sheet=None):
    try:
        # read_only=False required to preserve merged cell range metadata.
        # data_only=True returns cached formula results instead of formula strings.
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    except Exception as e:
        print(f"Error: Could not open Excel file: {e}", file=sys.stderr)
        sys.exit(1)

    filename = os.path.basename(file_path)
    sheet_names = wb.sheetnames

    if target_sheet and target_sheet not in sheet_names:
        print(
            f"Error: Sheet '{target_sheet}' not found. Available: {', '.join(sheet_names)}",
            file=sys.stderr,
        )
        wb.close()
        sys.exit(1)

    sheets_to_read = [target_sheet] if target_sheet else sheet_names
    parts = [
        f"# Excel: {filename}",
        f"*Sheets: {', '.join(sheet_names)}*",
        "",
        "> **Note:** Formula cells show cached values from the last save in Excel.",
        "> Cells with uncalculated formulas appear empty.",
    ]

    for name in sheets_to_read:
        parts.append(f"\n## Sheet: {name}\n")
        parts.append(sheet_to_markdown(wb[name]))

    wb.close()
    return "\n".join(parts)


def read_xls(file_path, target_sheet=None):
    try:
        import xlrd
    except ImportError:
        print(
            "Error: xlrd is not installed. Legacy .xls files require xlrd.\n"
            "Run: pip install xlrd",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        wb = xlrd.open_workbook(file_path)
    except Exception as e:
        print(f"Error: Could not open .xls file: {e}", file=sys.stderr)
        sys.exit(1)

    filename = os.path.basename(file_path)
    sheet_names = wb.sheet_names()

    if target_sheet and target_sheet not in sheet_names:
        print(
            f"Error: Sheet '{target_sheet}' not found. Available: {', '.join(sheet_names)}",
            file=sys.stderr,
        )
        sys.exit(1)

    sheets_to_read = [target_sheet] if target_sheet else sheet_names
    parts = [
        f"# Excel: {filename}",
        f"*Sheets: {', '.join(sheet_names)}*",
        "",
        "> **Note:** Legacy .xls format. Formula cells show cached values.",
    ]

    for name in sheets_to_read:
        ws = wb.sheet_by_name(name)
        parts.append(f"\n## Sheet: {name}\n")
        parts.append(xls_sheet_to_markdown(ws, wb.datemode))

    return "\n".join(parts)


def xls_sheet_to_markdown(ws, datemode):
    import xlrd

    if ws.nrows == 0:
        return "*Sheet is empty.*"

    def get_val(r, c):
        ctype = ws.cell_type(r, c)
        val = ws.cell_value(r, c)
        if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        if ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(val, datemode).strftime("%Y-%m-%d")
            except Exception:
                return str(val)
        if ctype == xlrd.XL_CELL_NUMBER:
            return int(val) if val == int(val) else val
        if ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if val else "FALSE"
        return val

    start = next(
        (r for r in range(ws.nrows) if any(ws.cell_type(r, c) not in (0, 6) for c in range(ws.ncols))),
        None,
    )
    if start is None:
        return "*Sheet is empty.*"

    headers = [format_value(get_val(start, c)) or f"Column {c + 1}" for c in range(ws.ncols)]
    while headers and not headers[-1].strip():
        headers.pop()

    ncols = len(headers)
    total_data = ws.nrows - start - 1
    truncated = total_data > MAX_ROWS
    end_row = min(ws.nrows, start + 1 + MAX_ROWS)

    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * ncols) + " |",
    ]

    for r in range(start + 1, end_row):
        cells = [format_value(get_val(r, c)) for c in range(ncols)]
        lines.append("| " + " | ".join(cells) + " |")

    result = "\n".join(lines)
    if truncated:
        result += f"\n\n> **Note:** Output truncated to {MAX_ROWS} rows. The sheet contains more data."
    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: read_excel.py <file_path> [sheet_name]", file=sys.stderr)
        sys.exit(1)

    file_path = normalize_path(sys.argv[1])
    target_sheet = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(file_path.lower())[1]

    if ext == ".xls":
        print(read_xls(file_path, target_sheet))
    else:
        print(read_xlsx(file_path, target_sheet))


if __name__ == "__main__":
    main()
