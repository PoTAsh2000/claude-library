#!/usr/bin/env python3
"""
Read an Excel file and output its contents as a markdown string to stdout.
Usage: read_excel.py <file_path> [sheet_name]
"""
import os
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_cell(cell):
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        return str(int(val)) if val == int(val) else str(val)
    return str(val).strip().replace("|", "\\|").replace("\n", " ")


def sheet_to_markdown(ws):
    rows = list(ws.iter_rows())

    # Skip leading empty rows
    start = next(
        (i for i, row in enumerate(rows) if any(c.value is not None for c in row)),
        None,
    )
    if start is None:
        return "*Sheet is empty.*"

    rows = rows[start:]
    headers = [format_cell(c) or f"Column {i + 1}" for i, c in enumerate(rows[0])]

    # Trim trailing empty columns
    while headers and not headers[-1].strip():
        headers.pop()

    if not headers:
        return "*Sheet is empty.*"

    ncols = len(headers)
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * ncols) + " |",
    ]

    for row in rows[1:]:
        cells = [format_cell(c) for c in row[:ncols]]
        while len(cells) < ncols:
            cells.append("")
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        print("Usage: read_excel.py <file_path> [sheet_name]", file=sys.stderr)
        sys.exit(1)

    file_path = sys.argv[1]
    target_sheet = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error: Could not open Excel file: {e}", file=sys.stderr)
        sys.exit(1)

    filename = os.path.basename(file_path)
    sheet_names = wb.sheetnames

    if target_sheet and target_sheet not in sheet_names:
        print(
            f"Error: Sheet '{target_sheet}' not found. Available: {', '.join(sheet_names)}",
            file=sys.stderr,
        )
        wb.close()
        sys.exit(1)

    sheets_to_read = [target_sheet] if target_sheet else sheet_names
    parts = [f"# Excel: {filename}", f"*Sheets: {', '.join(sheet_names)}*"]

    for name in sheets_to_read:
        parts.append(f"\n## Sheet: {name}\n")
        parts.append(sheet_to_markdown(wb[name]))

    wb.close()
    print("\n".join(parts))


if __name__ == "__main__":
    main()
